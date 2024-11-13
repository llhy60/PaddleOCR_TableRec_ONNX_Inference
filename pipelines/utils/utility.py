import io
import cv2
import math
import random
import numpy as np

from openpyxl import Workbook
from bs4 import BeautifulSoup
from PIL import Image, ImageDraw, ImageFont



def draw_ocr_box_txt(image,
                     boxes,
                     txts=None,
                     scores=None,
                     drop_score=0.5,
                     font_path="./doc/fonts/simfang.ttf"):
    h, w = image.height, image.width
    img_left = image.copy()
    img_right = np.ones((h, w, 3), dtype=np.uint8) * 255
    random.seed(0)

    draw_left = ImageDraw.Draw(img_left)
    if txts is None or len(txts) != len(boxes):
        txts = [None] * len(boxes)
    for idx, (box, txt) in enumerate(zip(boxes, txts)):
        if scores is not None and scores[idx] < drop_score:
            continue
        color = (random.randint(0, 255), random.randint(0, 255),
                 random.randint(0, 255))
        draw_left.polygon(box, fill=color)
        img_right_text = draw_box_txt_fine((w, h), box, txt, font_path)
        pts = np.array(box, np.int32).reshape((-1, 1, 2))
        cv2.polylines(img_right_text, [pts], True, color, 1)
        img_right = cv2.bitwise_and(img_right, img_right_text)
    img_left = Image.blend(image, img_left, 0.5)
    img_show = Image.new('RGB', (w * 2, h), (255, 255, 255))
    img_show.paste(img_left, (0, 0, w, h))
    img_show.paste(Image.fromarray(img_right), (w, 0, w * 2, h))
    return np.array(img_show)

def draw_box_txt_fine(img_size, box, txt, font_path="./doc/fonts/simfang.ttf"):
    box_height = int(
        math.sqrt((box[0][0] - box[3][0])**2 + (box[0][1] - box[3][1])**2))
    box_width = int(
        math.sqrt((box[0][0] - box[1][0])**2 + (box[0][1] - box[1][1])**2))

    if box_height > 2 * box_width and box_height > 30:
        img_text = Image.new('RGB', (box_height, box_width), (255, 255, 255))
        draw_text = ImageDraw.Draw(img_text)
        if txt:
            font = create_font(txt, (box_height, box_width), font_path)
            draw_text.text([0, 0], txt, fill=(0, 0, 0), font=font)
        img_text = img_text.transpose(Image.ROTATE_270)
    else:
        img_text = Image.new('RGB', (box_width, box_height), (255, 255, 255))
        draw_text = ImageDraw.Draw(img_text)
        if txt:
            font = create_font(txt, (box_width, box_height), font_path)
            draw_text.text([0, 0], txt, fill=(0, 0, 0), font=font)

    pts1 = np.float32(
        [[0, 0], [box_width, 0], [box_width, box_height], [0, box_height]])
    pts2 = np.array(box, dtype=np.float32)
    M = cv2.getPerspectiveTransform(pts1, pts2)

    img_text = np.array(img_text, dtype=np.uint8)
    img_right_text = cv2.warpPerspective(
        img_text,
        M,
        img_size,
        flags=cv2.INTER_NEAREST,
        borderMode=cv2.BORDER_CONSTANT,
        borderValue=(255, 255, 255))
    return img_right_text


def create_font(txt, sz, font_path="./doc/fonts/simfang.ttf"):
    font_size = int(sz[1] * 0.99)
    font = ImageFont.truetype(font_path, font_size, encoding="utf-8")
    length = font.getlength(txt)
    if length > sz[0]:
        font_size = int(font_size * sz[0] / length)
        font = ImageFont.truetype(font_path, font_size, encoding="utf-8")
    return font

def expand(pix, det_box, shape):
    x0, y0, x1, y1 = det_box
    h, w, c = shape
    tmp_x0 = x0 - pix
    tmp_x1 = x1 + pix
    tmp_y0 = y0 - pix
    tmp_y1 = y1 + pix
    x0_ = tmp_x0 if tmp_x0 >= 0 else 0
    x1_ = tmp_x1 if tmp_x1 <= w else w
    y0_ = tmp_y0 if tmp_y0 >= 0 else 0
    y1_ = tmp_y1 if tmp_y1 <= h else h
    return x0_, y0_, x1_, y1_

def handle_html2excel(pred_res):
    # try:
    pred_html = pred_res["html"]
    excel_data = to_excel(pred_html)
    # except Exception as e:  # noqa: F841
    #     pred_html = "<table><tr><th>无</th></tr><table>"
    #     excel_data = to_excel(pred_html)
    return excel_data

def to_excel(html_table):
    # 使用BeautifulSoup解析HTML表格
    soup = BeautifulSoup(html_table, "html.parser")
    table = soup.find("table")

    # 创建一个新的Excel工作簿和工作表
    wb = Workbook()
    ws = wb.active
    used_cells = set()
    # 遍历HTML表格的行并写入Excel工作表
    for row_idx, row in enumerate(table.find_all("tr"), 1):
        col_idx = 1
        for cell in row.find_all(["td", "th"]):
            cell_value = cell.get_text(strip=True)
            colspan = int(cell.get("colspan", 1))
            rowspan = int(cell.get("rowspan", 1))
            if (row_idx, col_idx) in used_cells:
                col_idx += colspan
                if col_idx > ws.max_column:
                    continue
            # 写入值，检查是否在合并单元格范围内
            if is_cell_in_merged_range(ws, row_idx, col_idx):
                # 获取下一个非合并的单元格位置
                next_row, next_col = get_next_non_merged_cell(
                    ws, row_idx, col_idx
                )
                # 将值写入下一个非合并的单元格
                ws.cell(row=next_row, column=next_col, value=cell_value)
                used_cells.add((next_row, next_col))
            else:
                ws.cell(row=row_idx, column=col_idx, value=cell_value)
                used_cells.add((row_idx, col_idx))

            # 合并单元格
            if colspan > 1 or rowspan > 1:
                ws.merge_cells(
                    start_row=row_idx,
                    start_column=col_idx,
                    end_row=row_idx + rowspan - 1,
                    end_column=col_idx + colspan - 1,
                )

            col_idx += colspan

    # 将工作簿保存到BytesIO对象
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)

    # 将指针移到文件开始位置，以便后续读取
    excel_buffer.seek(0)

    # 现在，excel_buffer包含Excel文件的二进制数据
    binary_excel_data = excel_buffer.getvalue()
    return binary_excel_data

# 检查单元格是否在任何合并范围内
def is_cell_in_merged_range(ws, row, column):
    cell = ws.cell(row=row, column=column)
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return True
    return False

# 获取下一个非合并的单元格位置
def get_next_non_merged_cell(ws, start_row, start_col):
    row, col = start_row, start_col
    while is_cell_in_merged_range(ws, row, col):
        col += 1
        if col > ws.max_column:
            col = 1
            row += 1
    return row, col